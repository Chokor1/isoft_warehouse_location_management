# Copyright (c) 2026, ISOFT LDA
# Author: Abbass Chokor
# For license information, please see license.txt
#
# Getting a warehouse into the app, and back out of it.
#
# Three things can be imported, in the order a warehouse is actually described: the zones
# it is divided into, the locations inside them, and what sits on each location. Each is
# a flat sheet with named columns, so a row that fails says which row and why rather than
# failing the file.
#
# Nothing is applied until it has all been checked. A half-imported warehouse is worse
# than an unimported one: you cannot tell what happened without reading every row, and
# re-running the file would double the quantities.

import frappe
from frappe import _
from frappe.utils import cint, cstr, flt

KINDS = ("zones", "locations", "stock")

COLUMNS = {
	"zones": ["warehouse", "zone_code", "zone_name", "sequence", "description"],
	"locations": [
		"warehouse", "location_code", "location_name", "zone", "location_type",
		"max_qty", "barcode", "description",
	],
	"stock": ["warehouse", "location_code", "item_code", "qty"],
}

# What each column is for, in the words someone filling the sheet in would use. This is
# not decoration: a spreadsheet that has to be explained in a separate email is a
# spreadsheet that comes back wrong.
NOTES = {
	"warehouse": "The exact warehouse name, as ERPNext spells it. Must be a leaf warehouse.",
	"zone_code": "Short code, unique inside the warehouse. Upper-cased automatically.",
	"zone_name": "What people call it. Left blank, the code is used.",
	"sequence": "Order the zones appear in. Lower comes first. Blank means 0.",
	"description": "Optional. A note for whoever reads the record later.",
	"location_code": "Short code, unique inside the warehouse — the label on the rack.",
	"location_name": "What people call it. Left blank, the code is used.",
	"zone": "The zone code this location sits in. It must already exist — import zones first.",
	"location_type": "One of: " + ", ".join(("Storage", "Pick Face", "Bulk", "Staging", "Quarantine"))
	+ ". Blank means Storage.",
	"max_qty": "How much this location can hold. Blank or 0 means no limit.",
	"barcode": "Optional. Scanned to identify the location.",
	"item_code": "The exact item code.",
	"qty": "How much of this item sits on this location. The difference comes from, or goes "
	"back to, unassigned stock — nothing leaves the warehouse.",
}

WIDTHS = {
	"warehouse": 32, "zone_code": 14, "zone_name": 22, "sequence": 10, "description": 40,
	"location_code": 16, "location_name": 24, "zone": 14, "location_type": 16,
	"max_qty": 12, "barcode": 16, "item_code": 22, "qty": 12,
}

REQUIRED = {
	"zones": ["warehouse", "zone_code"],
	"locations": ["warehouse", "location_code"],
	"stock": ["warehouse", "location_code", "item_code", "qty"],
}

SAMPLE = {
	"zones": ["01 - Loja Alvalade - ITEC", "FRONT", "Front Aisle", "1", "Fast movers by the till"],
	"locations": ["01 - Loja Alvalade - ITEC", "A-01", "Aisle A Shelf 1", "FRONT", "Pick Face",
	              "500", "7890001", "Reachable without a ladder"],
	"stock": ["01 - Loja Alvalade - ITEC", "A-01", "ITEM-0001", "12"],
}

LOCATION_TYPES = ("Storage", "Pick Face", "Bulk", "Staging", "Quarantine")


# ======================================================================
# helpers
# ======================================================================
def _guard():
	from isoft_warehouse_location_management.isoft_location_manager.page.isoft_location_manager.isoft_location_manager import (
		_require_preparer,
		_scope,
	)

	_require_preparer()
	return _scope(None)


def _kind(kind):
	if kind not in KINDS:
		frappe.throw(_("Unknown import type {0}.").format(kind))
	return kind


def _rows(rows):
	rows = frappe.parse_json(rows) if isinstance(rows, str) else rows
	if not isinstance(rows, list):
		frappe.throw(_("Expected a list of rows."))
	return rows


def _num(value, label, idx, problems, minimum=None, required=False):
	"""A number, or a clear complaint about what was actually in the cell."""
	raw = cstr(value).strip()
	if raw == "":
		if required:
			problems.append(_("row {0}: {1} is required").format(idx, label))
		return None
	# spreadsheets export 1 234,50 and 1,234.50 depending on the locale they were saved in
	cleaned = raw.replace(" ", "").replace(" ", "")
	if "," in cleaned and "." in cleaned:
		cleaned = cleaned.replace(",", "") if cleaned.rfind(".") > cleaned.rfind(",") else cleaned.replace(".", "").replace(",", ".")
	elif "," in cleaned:
		cleaned = cleaned.replace(",", ".")
	try:
		out = float(cleaned)
	except ValueError:
		problems.append(_("row {0}: {1} is not a number ({2})").format(idx, label, raw))
		return None
	if minimum is not None and out < minimum:
		problems.append(_("row {0}: {1} cannot be less than {2}").format(idx, label, minimum))
		return None
	return out


# ======================================================================
# templates and export
# ======================================================================
@frappe.whitelist()
def template(kind):
	"""An empty sheet with the right columns, and one row showing the shape."""
	_guard()
	kind = _kind(kind)
	return {"columns": COLUMNS[kind], "sample": SAMPLE[kind]}


@frappe.whitelist()
def export_rows(kind, warehouse=None):
	"""What is there now, in exactly the shape the importer accepts."""
	scope = _guard()
	kind = _kind(kind)

	def in_scope(wh):
		return scope is None or wh in scope

	out = []
	if kind == "zones":
		for z in frappe.get_all(
			"Warehouse Zone",
			filters={"warehouse": warehouse} if warehouse else {},
			fields=["warehouse", "zone_code", "zone_name", "sequence", "description"],
			order_by="warehouse asc, sequence asc, zone_code asc",
			limit_page_length=0,
		):
			if in_scope(z.warehouse):
				out.append([z.warehouse, z.zone_code, z.zone_name, cint(z.sequence), z.description or ""])

	elif kind == "locations":
		filters = {"is_unassigned": 0}
		if warehouse:
			filters["warehouse"] = warehouse
		for l in frappe.get_all(
			"Warehouse Location",
			filters=filters,
			fields=["warehouse", "location_code", "location_name", "zone", "location_type",
			        "max_qty", "barcode", "description"],
			order_by="warehouse asc, location_code asc",
			limit_page_length=0,
		):
			if not in_scope(l.warehouse):
				continue
			zone_code = frappe.db.get_value("Warehouse Zone", l.zone, "zone_code") if l.zone else ""
			out.append([l.warehouse, l.location_code, l.location_name or "", zone_code or "",
			            l.location_type or "", flt(l.max_qty) or "", l.barcode or "", l.description or ""])

	else:
		rows = frappe.get_all(
			"Location Stock",
			filters={"warehouse": warehouse} if warehouse else {},
			fields=["warehouse", "location", "item_code", "qty"],
			order_by="warehouse asc, location asc, item_code asc",
			limit_page_length=0,
		)
		codes = {}
		for r in rows:
			if not in_scope(r.warehouse) or flt(r.qty) <= 0:
				continue
			meta = codes.get(r.location)
			if meta is None:
				meta = frappe.db.get_value(
					"Warehouse Location", r.location, ["location_code", "is_unassigned"], as_dict=True
				) or frappe._dict()
				codes[r.location] = meta
			if meta.get("is_unassigned"):
				continue          # the remainder is derived; exporting it would re-import as a claim
			out.append([r.warehouse, meta.get("location_code"), r.item_code, flt(r.qty)])

	return {"columns": COLUMNS[kind], "rows": out}


# ======================================================================
# import
# ======================================================================
# Every row is checked before any row is written. `check` is the dry run the screen shows
# first; `apply` re-checks and only then commits, so a file cannot half-land because the
# warehouse changed between looking and pressing the button.


def _check_zones(rows, scope, pending=None):
	problems, plan = [], []
	seen = set()
	for i, r in enumerate(rows, start=1):
		wh = cstr(r.get("warehouse")).strip()
		code = cstr(r.get("zone_code")).strip().upper()
		if not wh or not code:
			problems.append(_("row {0}: warehouse and zone code are both required").format(i))
			continue
		if not frappe.db.exists("Warehouse", wh):
			problems.append(_("row {0}: no warehouse called {1}").format(i, wh))
			continue
		if frappe.get_cached_value("Warehouse", wh, "is_group"):
			problems.append(_("row {0}: {1} is a group warehouse — zones belong to a leaf").format(i, wh))
			continue
		if scope is not None and wh not in scope:
			problems.append(_("row {0}: {1} is not in your scope").format(i, wh))
			continue
		if (wh, code) in seen:
			problems.append(_("row {0}: {1} appears twice in this file").format(i, code))
			continue
		seen.add((wh, code))
		# a cell that could not be read is a reason to leave the row out of the plan, not
		# to quietly substitute a zero — the preview has to show what would really happen
		before = len(problems)
		_num(r.get("sequence"), _("sequence"), i, problems, minimum=0)
		if len(problems) > before:
			continue
		existing = frappe.db.get_value("Warehouse Zone", {"warehouse": wh, "zone_code": code}, "name")
		plan.append({
			"row": i, "action": "update" if existing else "create", "name": existing,
			"warehouse": wh, "zone_code": code,
			"zone_name": cstr(r.get("zone_name")).strip() or code.title(),
			"sequence": cint(flt(cstr(r.get("sequence")).strip() or 0)),
			"description": cstr(r.get("description")).strip(),
		})
	return plan, problems


def _check_locations(rows, scope, pending=None):
	from isoft_warehouse_location_management.isoft_location_manager.doctype.warehouse_location.warehouse_location import (
		UNASSIGNED_CODE,
	)

	problems, plan = [], []
	seen = set()
	for i, r in enumerate(rows, start=1):
		wh = cstr(r.get("warehouse")).strip()
		code = cstr(r.get("location_code")).strip().upper()
		if not wh or not code:
			problems.append(_("row {0}: warehouse and location code are both required").format(i))
			continue
		if not frappe.db.exists("Warehouse", wh):
			problems.append(_("row {0}: no warehouse called {1}").format(i, wh))
			continue
		if frappe.get_cached_value("Warehouse", wh, "is_group"):
			problems.append(_("row {0}: {1} is a group warehouse — locations belong to a leaf").format(i, wh))
			continue
		if scope is not None and wh not in scope:
			problems.append(_("row {0}: {1} is not in your scope").format(i, wh))
			continue
		if code == UNASSIGNED_CODE:
			problems.append(_("row {0}: {1} is the name of unassigned stock").format(i, code))
			continue
		if (wh, code) in seen:
			problems.append(_("row {0}: {1} appears twice in this file").format(i, code))
			continue
		seen.add((wh, code))

		zone_code = cstr(r.get("zone")).strip().upper()
		if zone_code:
			known = frappe.db.exists("Warehouse Zone", {"warehouse": wh, "zone_code": zone_code})
			# a zone the zones sheet of this same workbook is about to create counts as
			# existing: the sheets are applied in order, so by then it will
			coming = (wh, zone_code) in ((pending or {}).get("zones") or set())
			if not known and not coming:
				problems.append(
					_("row {0}: {1} has no zone {2} — import the zones first").format(i, wh, zone_code)
				)
				continue

		ltype = cstr(r.get("location_type")).strip()
		if ltype and ltype not in LOCATION_TYPES:
			problems.append(
				_("row {0}: {1} is not a location type ({2})").format(i, ltype, ", ".join(LOCATION_TYPES))
			)
			continue

		before = len(problems)
		max_qty = _num(r.get("max_qty"), _("capacity"), i, problems, minimum=0)
		if len(problems) > before:
			continue
		existing = frappe.db.get_value("Warehouse Location", {"warehouse": wh, "location_code": code}, "name")
		plan.append({
			"row": i, "action": "update" if existing else "create", "name": existing,
			"warehouse": wh, "location_code": code,
			"location_name": cstr(r.get("location_name")).strip() or code,
			"zone_code": zone_code, "location_type": ltype or "Storage",
			"max_qty": max_qty or 0, "barcode": cstr(r.get("barcode")).strip(),
			"description": cstr(r.get("description")).strip(),
		})
	return plan, problems


def _check_stock(rows, scope, pending=None):
	from isoft_warehouse_location_management.isoft_location_manager.doctype.location_stock.location_stock import (
		bin_qty,
		get_balance,
	)

	problems, plan = [], []
	# a file may name the same location and item twice; the last word wins, but say so
	claimed = {}
	for i, r in enumerate(rows, start=1):
		wh = cstr(r.get("warehouse")).strip()
		code = cstr(r.get("location_code")).strip().upper()
		item = cstr(r.get("item_code")).strip()
		if not wh or not code or not item:
			problems.append(_("row {0}: warehouse, location and item are all required").format(i))
			continue
		if scope is not None and wh not in scope:
			problems.append(_("row {0}: {1} is not in your scope").format(i, wh))
			continue
		if not frappe.db.exists("Item", item):
			problems.append(_("row {0}: no item called {1}").format(i, item))
			continue

		loc = frappe.db.get_value(
			"Warehouse Location", {"warehouse": wh, "location_code": code},
			["name", "is_active", "is_unassigned"], as_dict=True
		)
		coming = (wh, code) in ((pending or {}).get("locations") or set())
		if not loc and not coming:
			problems.append(_("row {0}: {1} has no location {2}").format(i, wh, code))
			continue
		if loc and loc.is_unassigned:
			problems.append(
				_("row {0}: unassigned stock is what is left over — it cannot be imported into").format(i)
			)
			continue
		if loc and not loc.is_active:
			problems.append(_("row {0}: location {1} is not active").format(i, code))
			continue

		qty = _num(r.get("qty"), _("qty"), i, problems, minimum=0, required=True)
		if qty is None:
			continue

		key = (loc.name if loc else (wh, code), item)
		if key in claimed:
			problems.append(
				_("row {0}: {1} on {2} is already set by row {3}").format(i, item, code, claimed[key])
			)
			continue
		claimed[key] = i

		plan.append({
			"row": i, "warehouse": wh, "location": loc.name if loc else None,
			"location_code": code, "item_code": item, "qty": qty,
			"was": get_balance(loc.name, item) if loc else 0,
		})

	# a location cannot be given more of an item than the warehouse holds
	wanted = {}
	for p in plan:
		wanted.setdefault((p["warehouse"], p["item_code"]), []).append(p)
	for (wh, item), group in wanted.items():
		on_hand = bin_qty(wh, item)
		# everything this file does not mention stays where it is
		touched = tuple(p["location"] for p in group if p["location"]) or ("",)
		others = flt(
			frappe.db.sql(
				"""select sum(qty) from `tabLocation Stock`
				   where warehouse=%s and item_code=%s and location not in %s""",
				(wh, item, touched),
			)[0][0]
		)
		asked = sum(flt(p["qty"]) for p in group)
		if asked + others > on_hand + 0.001:
			problems.append(
				_("{0} in {1}: the file places {2} but the warehouse holds {3}"
				  " ({4} is already on locations this file does not mention)").format(
					item, wh, asked, on_hand, others
				)
			)
	return plan, problems


CHECKERS = {"zones": _check_zones, "locations": _check_locations, "stock": _check_stock}


@frappe.whitelist()
def check(kind, rows):
	"""Dry run: what would happen, and everything wrong with the file."""
	scope = _guard()
	kind = _kind(kind)
	plan, problems = CHECKERS[kind](_rows(rows), scope, None)
	summary = {"create": 0, "update": 0, "set": 0}
	for p in plan:
		summary[p.get("action", "set")] = summary.get(p.get("action", "set"), 0) + 1
	return {
		"kind": kind,
		"ok": not problems,
		"rows": len(plan),
		"problems": problems,
		"summary": summary,
		"preview": plan[:20],
	}


@frappe.whitelist()
def apply(kind, rows):
	"""Write the file, or write nothing at all."""
	scope = _guard()
	kind = _kind(kind)
	plan, problems = CHECKERS[kind](_rows(rows), scope, None)
	if problems:
		frappe.throw(
			_("Nothing was imported — {0} problem(s) to fix first:<br>{1}").format(
				len(problems), "<br>".join(problems[:15])
			)
		)

	done = {"created": 0, "updated": 0, "placed": 0, "returned": 0}

	if kind == "zones":
		for p in plan:
			doc = frappe.get_doc("Warehouse Zone", p["name"]) if p["name"] else frappe.new_doc("Warehouse Zone")
			doc.update({
				"warehouse": p["warehouse"], "zone_code": p["zone_code"], "zone_name": p["zone_name"],
				"sequence": p["sequence"], "description": p["description"], "is_active": 1,
			})
			doc.flags.ignore_permissions = True
			doc.save(ignore_permissions=True)
			done["updated" if p["name"] else "created"] += 1

	elif kind == "locations":
		for p in plan:
			doc = (frappe.get_doc("Warehouse Location", p["name"]) if p["name"]
			       else frappe.new_doc("Warehouse Location"))
			# resolved now rather than at check time: the zone may have been created by an
			# earlier sheet of this same workbook
			zone = frappe.db.get_value(
				"Warehouse Zone", {"warehouse": p["warehouse"], "zone_code": p["zone_code"]}, "name"
			) if p["zone_code"] else None
			doc.update({
				"warehouse": p["warehouse"], "location_code": p["location_code"],
				"location_name": p["location_name"], "zone": zone,
				"location_type": p["location_type"], "max_qty": p["max_qty"],
				"barcode": p["barcode"], "description": p["description"], "is_active": 1,
			})
			doc.flags.ignore_permissions = True
			doc.save(ignore_permissions=True)
			done["updated" if p["name"] else "created"] += 1

	else:
		from isoft_warehouse_location_management.isoft_location_manager.page.isoft_location_manager.isoft_location_manager import (
			set_location_qty,
		)

		for p in plan:
			# likewise: the location may have been created a moment ago by the locations
			# sheet of this same workbook
			location = p["location"] or frappe.db.get_value(
				"Warehouse Location",
				{"warehouse": p["warehouse"], "location_code": p["location_code"]}, "name"
			)
			if not location:
				frappe.throw(
					_("Row {0}: {1} has no location {2}.").format(
						p["row"], p["warehouse"], p["location_code"]
					)
				)
			# the board's own path: the difference comes from, or goes back to, unassigned
			# stock, and the movement is recorded either way
			r = set_location_qty(p["warehouse"], p["item_code"], location, p["qty"])
			moved = flt((r or {}).get("changed"))
			if moved > 0:
				done["placed"] += moved
			elif moved < 0:
				done["returned"] += -moved
			done["updated"] += 1

	return {"ok": True, "kind": kind, "rows": len(plan), **done}


# ======================================================================
# workbooks
# ======================================================================
# A sheet somebody has to be told how to fill in is a sheet that comes back wrong, so the
# workbook explains itself: the header says what is required, the columns are wide enough
# to read, the cells that have a fixed set of answers offer them as a dropdown, and a
# second sheet spells out every column in words.

HEADER_FILL = "FF0F766E"
REQUIRED_FILL = "FFE6FFFA"


def _sheet_title(kind):
	return {"zones": "Zones", "locations": "Locations", "stock": "Stock on locations"}[kind]


def _sheet_into(wb, kind, rows, with_sample=False):
	from openpyxl import Workbook
	from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
	from openpyxl.utils import get_column_letter
	from openpyxl.worksheet.datavalidation import DataValidation

	columns = COLUMNS[kind]
	required = REQUIRED[kind]

	if wb is None:
		wb = Workbook()
		ws = wb.active
		ws.title = _sheet_title(kind)
	else:
		ws = wb.create_sheet(_sheet_title(kind))

	head_font = Font(bold=True, color="FFFFFFFF", size=11)
	head_fill = PatternFill("solid", fgColor=HEADER_FILL)
	thin = Side(style="thin", color="FFD8E0E6")

	for c, col in enumerate(columns, start=1):
		cell = ws.cell(row=1, column=c)
		# the header carries the rule, so nobody has to look it up
		cell.value = col + (" *" if col in required else "")
		cell.font = head_font
		cell.fill = head_fill
		cell.alignment = Alignment(vertical="center", horizontal="left")
		cell.border = Border(bottom=thin)
		ws.column_dimensions[get_column_letter(c)].width = WIDTHS.get(col, 18)
	ws.row_dimensions[1].height = 22
	ws.freeze_panes = "A2"
	ws.auto_filter.ref = "A1:%s1" % get_column_letter(len(columns))

	start = 2
	if with_sample:
		for c, value in enumerate(SAMPLE[kind], start=1):
			cell = ws.cell(row=2, column=c, value=value)
			cell.font = Font(italic=True, color="FF8D99A6")
		start = 3

	for r, row in enumerate(rows or [], start=start):
		for c, value in enumerate(row, start=1):
			ws.cell(row=r, column=c, value=value)

	# quantities and capacities are numbers, and should look like numbers
	for c, col in enumerate(columns, start=1):
		if col in ("qty", "max_qty", "sequence"):
			for r in range(2, ws.max_row + 1):
				ws.cell(row=r, column=c).number_format = "#,##0.###"

	# the columns with a fixed set of answers offer them, rather than being guessed at
	limit = max(ws.max_row + 200, 500)
	if "location_type" in columns:
		c = get_column_letter(columns.index("location_type") + 1)
		dv = DataValidation(
			type="list", formula1='"%s"' % ",".join(LOCATION_TYPES), allow_blank=True,
			showErrorMessage=True, errorTitle="Not a location type",
			error="Choose one of: " + ", ".join(LOCATION_TYPES),
		)
		ws.add_data_validation(dv)
		dv.add("%s2:%s%d" % (c, c, limit))
	for col in ("qty", "max_qty"):
		if col in columns:
			c = get_column_letter(columns.index(col) + 1)
			dv = DataValidation(
				type="decimal", operator="greaterThanOrEqual", formula1=0, allow_blank=(col != "qty"),
				showErrorMessage=True, errorTitle="Not a quantity",
				error="This must be a number, and cannot be negative.",
			)
			ws.add_data_validation(dv)
			dv.add("%s2:%s%d" % (c, c, limit))

	return wb


def _help_sheet(wb, kind, suffix=False):
	from openpyxl.styles import Alignment, Font, PatternFill

	title = "How to fill in %s" % _sheet_title(kind) if suffix else "How to fill this in"
	ws = wb.create_sheet(title[:31])
	ws.column_dimensions["A"].width = 20
	ws.column_dimensions["B"].width = 12
	ws.column_dimensions["C"].width = 86

	ws["A1"] = _sheet_title(kind)
	ws["A1"].font = Font(bold=True, size=14)
	ws["A2"] = {
		"zones": "The parts a warehouse is divided into. Import these before locations, "
		"because a location can name one.",
		"locations": "The shelves, racks and bins themselves. A zone named here must already exist.",
		"stock": "What sits on each location. The difference comes from, or goes back to, "
		"unassigned stock — nothing leaves the warehouse.",
	}[kind]
	ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
	ws.merge_cells("A2:C2")
	ws.row_dimensions[2].height = 30

	for c, title in enumerate(("Column", "Required", "What it means"), start=1):
		cell = ws.cell(row=4, column=c, value=title)
		cell.font = Font(bold=True, color="FFFFFFFF")
		cell.fill = PatternFill("solid", fgColor=HEADER_FILL)

	for r, col in enumerate(COLUMNS[kind], start=5):
		ws.cell(row=r, column=1, value=col).font = Font(bold=True)
		ws.cell(row=r, column=2, value="yes" if col in REQUIRED[kind] else "")
		cell = ws.cell(row=r, column=3, value=NOTES.get(col, ""))
		cell.alignment = Alignment(wrap_text=True, vertical="top")

	last = 5 + len(COLUMNS[kind]) + 1
	ws.cell(row=last, column=1, value="Before you import").font = Font(bold=True)
	ws.cell(
		row=last + 1, column=1,
		value="The whole file is checked before anything is written. If any row is wrong, "
		"nothing is imported at all — fix the rows it lists and send the same file again.",
	).alignment = Alignment(wrap_text=True, vertical="top")
	ws.merge_cells(start_row=last + 1, start_column=1, end_row=last + 1, end_column=3)
	ws.row_dimensions[last + 1].height = 30
	return ws


def _respond(wb, filename):
	from io import BytesIO

	buf = BytesIO()
	wb.save(buf)
	frappe.response["filename"] = filename
	frappe.response["filecontent"] = buf.getvalue()
	frappe.response["type"] = "binary"


def _kinds(kinds):
	"""One or several kinds, in the order a warehouse has to be described."""
	if isinstance(kinds, str):
		kinds = frappe.parse_json(kinds) if kinds.strip().startswith("[") else [kinds]
	wanted = [k for k in KINDS if k in set(kinds or [])]
	if not wanted:
		frappe.throw(_("Choose at least one thing to include."))
	return wanted


def _multi_workbook(kinds, rows_by_kind, with_sample=False):
	"""A sheet per kind, in dependency order, with the help sheets after them."""
	wb = None
	for kind in kinds:
		wb = _sheet_into(wb, kind, rows_by_kind.get(kind) or [], with_sample)
	for kind in kinds:
		_help_sheet(wb, kind, suffix=len(kinds) > 1)
	return wb


@frappe.whitelist()
def download_template(kinds="locations"):
	"""Empty sheets with the right columns, an example row each, and how to fill them in."""
	_guard()
	kinds = _kinds(kinds)
	wb = _multi_workbook(kinds, {}, with_sample=True)
	name = "ilm-template.xlsx" if len(kinds) > 1 else "ilm-%s-template.xlsx" % kinds[0]
	_respond(wb, name)


@frappe.whitelist()
def download_export(kinds="locations", warehouse=None):
	"""What is there now, in exactly the workbook the importer accepts."""
	_guard()
	kinds = _kinds(kinds)
	rows_by_kind = {k: export_rows(k, warehouse)["rows"] for k in kinds}
	wb = _multi_workbook(kinds, rows_by_kind)
	name = "ilm-export.xlsx" if len(kinds) > 1 else "ilm-%s.xlsx" % kinds[0]
	_respond(wb, name)


def _kind_of(header):
	"""Which of the three a sheet is, judged by its own columns.

	The distinguishing column is enough and does not depend on order: only stock names an
	item, only zones name a zone code of their own, and what is left with a location code
	is the locations sheet.
	"""
	cols = set(header)
	if "item_code" in cols:
		return "stock"
	if "zone_code" in cols:
		return "zones"
	if "location_code" in cols:
		return "locations"
	return None


def _header_key(h):
	return cstr(h).strip().lower().replace("*", "").replace(" ", "_").strip("_")


def _table_to_rows(table):
	table = [r for r in (table or []) if any(cstr(c).strip() for c in r)]
	if not table:
		return None, []
	header = [_header_key(h) for h in table[0]]
	kind = _kind_of(header)
	if not kind:
		return None, []
	rows = []
	for line in table[1:]:
		row = {}
		for i, h in enumerate(header):
			if h:
				row[h] = line[i] if i < len(line) else ""
		rows.append(row)
	return kind, rows


@frappe.whitelist()
def read_upload(content, filename=None):
	"""Turn an uploaded workbook into sheets of rows, each one identified by its header.

	A workbook may hold zones, locations and stock at once — that is what the export
	produces — so every sheet is read and identified on its own. Sheets that are not one
	of the three are skipped rather than complained about: a working spreadsheet usually
	has a scratch tab in it.

	`.csv` is accepted too, because somebody will always send one, and refusing it
	teaches them nothing.
	"""
	_guard()
	import base64

	raw = content
	if isinstance(raw, str):
		if raw[:5] == "data:" and "," in raw[:128]:
			raw = raw.split(",", 1)[1]
		raw = base64.b64decode(raw)

	tables = []
	if (filename or "").lower().endswith(".csv"):
		import csv
		import io

		text = raw.decode("utf-8-sig", errors="replace")
		tables.append([r for r in csv.reader(io.StringIO(text))])
	else:
		import io

		from openpyxl import load_workbook

		wb = load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
		for ws in wb.worksheets:
			tables.append([[c for c in row] for row in ws.iter_rows(values_only=True)])

	sheets = []
	for table in tables:
		kind, rows = _table_to_rows(table)
		if kind and rows:
			sheets.append({"kind": kind, "rows": rows, "count": len(rows)})

	if not sheets:
		frappe.throw(
			_("Nothing in that file looks like zones, locations or stock. "
			  "The first row of a sheet has to be the column names — start from a template.")
		)

	# always in the order a warehouse has to be described
	sheets.sort(key=lambda x: KINDS.index(x["kind"]))
	return {"sheets": sheets, "count": sum(s["count"] for s in sheets)}


@frappe.whitelist()
def check_sheets(sheets):
	"""Dry run every sheet, in order, and report them together."""
	scope = _guard()
	sheets = frappe.parse_json(sheets) if isinstance(sheets, str) else sheets
	sheets = sorted(sheets or [], key=lambda x: KINDS.index(_kind(x.get("kind"))))
	out, ok_all = [], True
	# what the sheets before this one will have created by the time it is applied
	pending = {"zones": set(), "locations": set()}
	for sheet in sheets:
		kind = _kind(sheet.get("kind"))
		plan, problems = CHECKERS[kind](sheet.get("rows") or [], scope, pending)
		if kind == "zones":
			pending["zones"] |= {(p["warehouse"], p["zone_code"]) for p in plan}
		elif kind == "locations":
			pending["locations"] |= {(p["warehouse"], p["location_code"]) for p in plan}
		summary = {"create": 0, "update": 0, "set": 0}
		for p in plan:
			key = p.get("action", "set")
			summary[key] = summary.get(key, 0) + 1
		ok_all = ok_all and not problems
		out.append({
			"kind": kind, "title": _sheet_title(kind), "rows": len(plan),
			"read": len(sheet.get("rows") or []), "problems": problems, "summary": summary,
		})
	return {"ok": ok_all, "sheets": out,
	        "rows": sum(s["rows"] for s in out),
	        "problems": sum(len(s["problems"]) for s in out)}


@frappe.whitelist()
def apply_sheets(sheets):
	"""Write every sheet, or write nothing at all.

	Checked as a whole first: locations that name a zone in the same file are only valid
	once that zone exists, so the sheets are applied in order — but a problem anywhere
	stops all of it, including the sheets that would have been fine.
	"""
	_guard()
	sheets = frappe.parse_json(sheets) if isinstance(sheets, str) else sheets
	sheets = sorted(sheets or [], key=lambda x: KINDS.index(_kind(x.get("kind"))))

	report = check_sheets(sheets)
	if not report["ok"]:
		lines = []
		for s in report["sheets"]:
			lines += ["<b>%s</b>: %s" % (s["title"], p) for p in s["problems"][:10]]
		frappe.throw(
			_("Nothing was imported — {0} problem(s) to fix first:<br>{1}").format(
				report["problems"], "<br>".join(lines[:15])
			)
		)

	done = []
	for sheet in sheets:
		kind = _kind(sheet.get("kind"))
		r = apply(kind, frappe.as_json(sheet.get("rows") or []))
		r["title"] = _sheet_title(kind)
		done.append(r)
	return {"ok": True, "sheets": done}
